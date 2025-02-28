import openpyxl # type: ignore
import pandas as pd # type: ignore
import logging
import os



# --------------------------------------------------------------------GET LATEST EXCEL FILE  ----------------------------
Excel_file_folder = r"C:\Users\hss930284\Tata Technologies\MBSE Team - SAARCONN - SAARCONN\Eliminating_SystemDesk\tests\Harshit_validation_27_02\COMBINED_AUTOMATION\Input_Excel"

file2_list = [os.path.join(Excel_file_folder, f) for f in os.listdir(Excel_file_folder) if f.endswith(".xlsx")]
if not file2_list:
   raise FileNotFoundError(f"No Excel files found in {Excel_file_folder}")

Excel_file_path = max(file2_list, key=os.path.getmtime)
# --------------------------------------------------------------------GET LATEST EXCEL FILE  ----------------------------



# --------------------------------------------------------------------GET FILE NAME ----------------------------

def get_latest_excel_name(folder_path):
   file_list = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx")]
   if not file_list:
       raise FileNotFoundError(f"No Excel files found in {folder_path}")
   latest_file = max(file_list, key=os.path.getmtime)
   return os.path.splitext(os.path.basename(latest_file))[0]
Excel_file_folder = Excel_file_folder
Excelfile_name = get_latest_excel_name(Excel_file_folder)
# --------------------------------------------------------------------GET FILE NAME ----------------------------


# --------------------------------------------------------------------GET GENERATED ARXML FILE PATH ----------------------------

print("The latest Excel file name is:", Excelfile_name)
Arxml_directory = r"C:\Users\hss930284\Tata Technologies\MBSE Team - SAARCONN - SAARCONN\Eliminating_SystemDesk\tests\Harshit_validation_27_02\COMBINED_AUTOMATION\Intermidiate_Outputs\Generated_ARXML"
os.makedirs(Arxml_directory, exist_ok=True)        
arxml_file_path = os.path.join(Arxml_directory, f"{Excelfile_name}.arxml")
# --------------------------------------------------------------------GET GENERATED ARXML FILE PATH ----------------------------



class ExcelReader:
    def __init__(self):
        """
        Initialize the ExcelReader without a file path.
        The file path will be provided by the user later.
        """
        
        
        self.file_path = Excel_file_path
        self.workbook = None
        self.xls = None
        self.project_info = None
        self.swc_info = None
        self.ib_data = None
        self.ports = None
        self.adt_primitive = None
        self.adt_composite = None
        self.idt = None

    def get_file_path_from_user(self):
        """
        Prompts the user to input the path of the Excel file.
        Validates the file path and sets it as the instance attribute.
        """
        while True:
            self.file_path = Excel_file_path
            if os.path.isfile(self.file_path):
                break
            else:
                print("Error: Invalid file path. Please try again.")

    def read_user_defined_excel(self):
        """
        Loads the workbook and assigns each worksheet to a corresponding variable for easy access.
        Returns the workbook and ExcelFile object.
        """
        if not self.file_path:
            raise ValueError("File path is not set. Call `get_file_path_from_user()` first.")

        try:
            # Load the workbook and ExcelFile object
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.xls = pd.ExcelFile(self.file_path)
        except (FileNotFoundError, PermissionError, openpyxl.utils.exceptions.InvalidFileException) as e:
            logging.error(f"Error: Unable to load workbook due to {e}")
            raise Exception(f"Unable to load workbook: {e}")

        # Assign worksheets to class attributes
        worksheet_names = [
            "project_info", 
            "swc_info", 
            "ib_data", 
            "ports", 
            "adt_primitive", 
            "adt_composite",
            "idt"
        ]
        try:
            self.project_info, self.swc_info, self.ib_data, self.ports, self.adt_primitive, self.adt_composite, self.idt = [
                self.workbook.worksheets[i] for i in range(len(worksheet_names))
            ]
        except IndexError as e:
            logging.error(f"Error: Unable to access worksheets due to {e}")
            raise Exception(f"Unable to access worksheets: {e}")

        return self.workbook, self.xls

    def column_letter_to_index(self, column_letter):
        """
        Converts Excel column letter (e.g., 'A', 'Z', 'AA', 'AZ') to a 1-based column index.
        Args:
            column_letter: A string representing the column letter(s) (e.g., 'A', 'Z', 'AA', 'AZ').
        Returns:
            int: The 1-based index of the column.
        """
        column_index = 0
        for char in column_letter:
            column_index = column_index * 26 + (ord(char.upper()) - ord('A') + 1)
        return column_index
    
    def read_columns(self, current_sheet, first_col, last_col, start_row=2, end_row=None):
        """
        Reads specified columns from the given sheet and returns a list of filtered column data.
        This version ensures that merged cells are handled properly, propagating their value only
        once for each merged block when reading columns individually.
        Parameters:
            current_sheet: The sheet to read from.
            first_col: The first column letter to read.
            last_col: The last column letter to read.
            start_row: The row to start reading from (default is 2).
            end_row: The row to stop reading at (default is None, which means the last row).
        Returns:
            A list of lists containing the filtered column data (no `None` values).
        """
        column_data = []
        try:
            if end_row is None:
                end_row = current_sheet.max_row
        except AttributeError as e:
            logging.error(f"Error: Unable to access max_row attribute due to {e}")
            raise Exception(f"Unable to access max_row: {e}")

        try:
            first_col_index = self.column_letter_to_index(first_col)
            last_col_index = self.column_letter_to_index(last_col)
            num_columns = last_col_index - first_col_index + 1
        except Exception as e:
            logging.error(f"Error: Unable to calculate column indices due to {e}")
            raise Exception(f"Unable to calculate column indices: {e}")

        valid_columns = [[] for _ in range(num_columns)]

        # To track the last seen value for each column (used for propagating merged values)
        last_seen_values = [None] * num_columns

        # Iterate through each row to read the data
        for row in range(start_row, end_row + 1):
            for col in range(num_columns):
                cell_value = current_sheet.cell(row=row, column=first_col_index + col).value

                # Check if the current cell is part of a merged range (not the first cell in the range)
                if cell_value is None:
                    # For merged cells, propagate the value from the previous row (only once for the first row in the merged range)
                    cell_value = last_seen_values[col]
                else:
                    # Update the last seen value (for merged cells or new data cells)
                    last_seen_values[col] = cell_value

                valid_columns[col].append(cell_value)

        # Filter out None values from each column
        filtered_columns = [list(filter(lambda x: x is not None, column)) for column in valid_columns]
        return valid_columns

