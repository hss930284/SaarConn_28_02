import openpyxl

import re

import os

from datetime import datetime

import openpyxl.utils

from openpyxl.utils import get_column_letter

# ANSI escape codes for color output

RED = "\033[91m"  # Critical (Red)

YELLOW = "\033[93m"  # Warning (Yellow)

BLUE = "\033[94m"  # Info (Blue)

GREEN = "\033[92m"  # Success (Green)

RESET = "\033[0m"  # Reset color to default

# Dictionary to store validation errors

errors = {

   "Critical": [],

   "Warning": [],

   "Info": []

}

def validate_excel(file_path):
    """ Validates the Excel file based on provided rules. """
    try:
        wb = openpyxl.load_workbook(file_path)

        ### 🔵 Empty Cell Validation (`excel_rule_1`) ###

        """"
        excel_rule_1: Non-Empty Cells with Specific Exceptions

        Rule Definition:
            All data cells within the specified Excel sheets must contain a value. 
            The first row of each sheet is designated as the header row and is therefore excluded from this validation.
            Data validation will commence from the second row onwards.
            
        Exceptions and Warnings:

            The following columns and cells are exempt from the non-empty cell requirement, and any deviations from the specified conditions 
            should be flagged as warnings:
                1. "swc_info" Sheet:
                    * Columns D and I: These columns may contain empty cells.
                    * Column M: The content of this column is contingent upon the value in column L, as detailed below:
                    * For column L values of 'AsynchronousServerCallReturnsEvent', 'BackgroundEvent', 'InitEvent', 'InternalTriggerOccurredEvent',
                        'OsTaskExecutionEvent', 'SwcModeManagerErrorEvent' and 'TransformerHardErrorEvent', column M must be empty.
                    * For column L values of 'DataReceivedEvent', 'DataReceiveErrorEvent', 'DataSendCompletedEvent', 'DataWriteCompletedEvent',
                        'ExternalTriggerOccurredEvent', 'ModeSwitchedAckEvent', 'OperationInvokedEvent', and 'SwcModeSwitchEvent', column M must
                        contain a port name. This port name must correspond to a value found in column C of the "ports" sheet.
                            * Furthermore, the corresponding row in the "ports" sheet must satisfy the following criteria:
                            * For 'DataReceivedEvent' and 'DataReceiveErrorEvent' the corresponding "B" column value must be "ReceiverPort" and
                                "D" column value must be either "SenderReceiverInterface" or "NvDataInterface".
                            * For 'DataSendCompletedEvent' and 'DataWriteCompletedEvent' the corresponding "B" column value must be "SenderPort"
                                and "D" column value must be either "SenderReceiverInterface" or "NvDataInterface".
                            * For 'ExternalTriggerOccurredEvent' the corresponding "B" column value must be "ReceiverPort" and "D" column value 
                                must be "TriggerInterface".
                            * For 'ModeSwitchedAckEvent' and 'OperationInvokedEvent' the corresponding "B" column value must be "SenderPort" and
                                "D" column value must be "ModeSwitchInterface" and "ClientServerInterface" respectively.
                            * For 'SwcModeSwitchEvent' the corresponding "B" column value must be "ReceiverPort" and "D" column value must be
                                "ModeSwitchInterface".
                            * Example: If column M contains 'rport1', and 'rport1' is located in cell C7 of the "ports" sheet, then cells B7 and D7 of
                                the "ports" sheet must contain the respective corresponding values as described above.
                    * For column L value of 'TimingEvent' column M must contain a numeric time value (e.g., 1.0, 0.87).
                2. "ib_data" Sheet:
                    * Column E: This column may contain empty cells.
                    * Column M: If corresponding column B value contains either "PerInstanceMemory" or "ArTypedPerInstanceMemory", then column M
                                must be empty.
                3. "ports" Sheet:
                    * Columns J, K, and L: These columns may contain empty cells.
                4. "adt_primitive" Sheet:
                    * If column E contains the value 'IDENTICAL', then corresponding columns F and G values must be empty. Check column E also for None values.
                5. "idt" Sheet:
                    * If column B contains the value 'PRIMITIVE', then corresponding column D value must be empty. Check column B also for None values.

        """
        # Define sheets and exception columns
        empty_check_sheets = {
        "swc_info": ["D", "I", "M"],
        "ib_data": ["E"],
        "ports": ["J", "K", "L"],
        "adt_primitive": ["E", "F", "G"],
        "idt": ["B", "D"]
        }
        column_limits = {
        "swc_info": "M",
        "ib_data": "I",
        "ports": "L",
        "adt_primitive": "M",
        "adt_composite": "F",
        "idt": "E"
        }
        # Load sheets
        swc_info = wb["swc_info"] if "swc_info" in wb.sheetnames else None
        ports = wb["ports"] if "ports" in wb.sheetnames else None
        ib_data = wb["ib_data"] if "ib_data" in wb.sheetnames else None
        adt_primitive = wb["adt_primitive"] if "adt_primitive" in wb.sheetnames else None
        idt = wb["idt"] if "idt" in wb.sheetnames else None

        # Function to get merged cell mappings
        def get_merged_cells(sheet):
            merged_ranges = {}
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                first_cell = f"{get_column_letter(min_col)}{min_row}"
                merged_cells = [
                    f"{get_column_letter(col)}{row}"
                    for col in range(min_col, max_col + 1)
                    for row in range(min_row, max_row + 1)
                    if f"{get_column_letter(col)}{row}" != first_cell
                ]
                merged_ranges[first_cell] = merged_cells
            return merged_ranges
        # 🔹 1️⃣ General non-empty validation
        for sheet_name, exception_columns in empty_check_sheets.items():
            sheet = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if sheet:
                merged_ranges = get_merged_cells(sheet)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    for col_idx, cell in enumerate(row):
                        column_letter = get_column_letter(col_idx + 1)
                        cell_ref = f"{column_letter}{row_idx}"
                        # Skip merged cells (except first cell)
                        if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                            errors["Info"].append(f"[{sheet_name}] Merged cell {cell_ref} is expected to be empty")
                            continue  
                        # Skip exception columns
                        if column_letter in exception_columns:
                            continue  
                        # Check for missing value
                        if cell.value in [None, ""]:
                            errors["Critical"].append(f"[{sheet_name}] Missing value at {cell_ref}")
        # 🔹 2️⃣ Get ports mapping
        def get_ports_mapping():
            if not ports:
                return {}
            ports_map = {}
            for row in ports.iter_rows(min_row=2, values_only=True):
                if len(row) > 3:
                    port_name, b_value, d_value = row[2], row[1], row[3]  # Columns C, B, D
                    if port_name:
                        ports_map[port_name] = (b_value, d_value)
            return ports_map
        ports_mapping = get_ports_mapping()
        # 🔹 3️⃣ Column M validation rules for swc_info
        m_validation_rules = {
        "AsynchronousServerCallReturnsEvent": lambda m: m is None,
        "BackgroundEvent": lambda m: m is None,
        "DataReceivedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("ReceiverPort", "SenderReceiverInterface"), ("ReceiverPort", "NvDataInterface") ],
        "DataReceiveErrorEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("ReceiverPort", "SenderReceiverInterface"), ("ReceiverPort", "NvDataInterface") ],
        "DataSendCompletedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("SenderPort", "SenderReceiverInterface"), ("SenderPort", "NvDataInterface")],
        "DataWriteCompletedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("SenderPort", "SenderReceiverInterface"), ("SenderPort", "NvDataInterface") ],
        "ExternalTriggerOccurredEvent": lambda m: ports_mapping.get(m, (None, None)) == ("ReceiverPort", "TriggerInterface"),
        "InitEvent": lambda m: m is None,
        "InternalTriggerOccurredEvent": lambda m: m is None,
        "ModeSwitchedAckEvent": lambda m: ports_mapping.get(m, (None, None)) == ("SenderPort", "ModeSwitchInterface"),
        "OperationInvokedEvent": lambda m: ports_mapping.get(m, (None, None)) == ("SenderPort", "ClientServerInterface"),
        "OsTaskExecutionEvent": lambda m: m is None,
        "SwcModeManagerErrorEvent": lambda m: m is None,
        "SwcModeSwitchEvent": lambda m: ports_mapping.get(m, (None, None)) == ("ReceiverPort", "ModeSwitchInterface"),
        "TimingEvent": lambda m: isinstance(m, (int, float)),
        "TransformerHardErrorEvent": lambda m: m is None,
        }
        # Validate column M in swc_info
        if swc_info:
            for row_idx, row in enumerate(swc_info.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > 12:
                    event_type = row[11]  # Column L
                    m_value = row[12]  # Column M
                    cell_ref = f"M{row_idx}"
                    if event_type in m_validation_rules and not m_validation_rules[event_type](m_value):
                        errors["Critical"].append(f"[swc_info] Invalid value at {cell_ref} for event type '{event_type}'")
            # 🔹 4️⃣ Validate column M in ib_data
        if ib_data:
            for row_idx, row in enumerate(ib_data.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > 12:
                    col_b_value = row[1]  # Column B
                    col_m_value = row[12]  # Column M
                    cell_ref = f"M{row_idx}"
                    if col_b_value in ["PerInstanceMemory", "ArTypedPerInstanceMemory"] and col_m_value is not None:
                        errors["Critical"].append(f"[ib_data] Column M must be empty at {cell_ref} when Column B is '{col_b_value}'")

        # 🔹 5️⃣ Validate column E, F & G in adt_primitive
        if adt_primitive:
            merged_ranges = get_merged_cells(adt_primitive)  # Get merged ranges
            for row_idx, row in enumerate(adt_primitive.iter_rows(min_row=2), start=2):
                for col_idx in [4, 5, 6]:  # Columns E, F, G (0-based index: 4, 5, 6)
                    column_letter = get_column_letter(col_idx + 1)
                    cell_ref = f"{column_letter}{row_idx}"
                    cell = row[col_idx]
                    # Skip merged cells (except first cell)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[adt_primitive] Merged cell {cell_ref} is expected to be empty")
                        continue  
                    # Validate based on rules
                    if column_letter == "E" and cell.value in [None, ""]:
                        errors["Critical"].append(f"[adt_primitive] Column E must not be empty at {cell_ref}")
                    elif column_letter in ["F", "G"] and row[4].value == "IDENTICAL" and cell.value not in [None, ""]:
                        errors["Critical"].append(f"[adt_primitive] Column {column_letter} must be empty at {cell_ref} when Column E is 'IDENTICAL'")
                                
        # 🔹 6️⃣Validate column B & D in idt
        if idt:
            merged_ranges = get_merged_cells(idt)  # Get merged ranges
            for row_idx, row in enumerate(idt.iter_rows(min_row=2), start=2):
                for col_idx in [1, 3]:  # Columns B, D (0-based index: 1, 3)
                    column_letter = get_column_letter(col_idx + 1)
                    cell_ref = f"{column_letter}{row_idx}"
                    cell = row[col_idx]
                    # Skip merged cells (except first cell)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[idt] Merged cell {cell_ref} is expected to be empty")
                        continue  
                    # Validate based on rules
                    if column_letter == "B" and cell.value in [None, ""]:
                        errors["Critical"].append(f"[idt] Column B must not be empty at {cell_ref}")
                    elif column_letter == "D" and row[1].value == "PRIMITIVE" and cell.value not in [None, ""]:
                        errors["Critical"].append(f"[idt] Column D must be empty at {cell_ref} when Column B is 'PRIMITIVE'")

        ### 🟢 Naming Convention Rule ('excel_rule_2') ###

        """
        excel_rule_2 : Naming convention
            . this rule is applicable to following user given values             
                in "swc_info" column "C",  column "D", column "E", column "H", column "I", and column "K"
                in "ib_data" column "C"
                in "ports" column "C",  column "E", column "F", and column "G"
                in "adt_primitive" column "B",  column "D", column "G", and column "I"
                in "adt_composite" column "C" and  column "D", , except for numbers in D column if corresponding B value is ARRAY or ARRAY_FIXED or ARRAY_VARIABLE
                    handle combinely for IDT as well
                in "idt" column "C",  and column "D" , , except for numbers in D column
            . first row of every sheet will be the header so the data for validation should be consider from second row of each above mentioned excel sheets.
            . the rule is 'the name can have small and capital alphabetical letters and numbers from 0 to 9 and no special characters except _ '            
            . the name can start with only alphabetical which can be either capital or small letters
        """
        
        # ✅ Define the sheets and columns to check naming conventions
        naming_sheets = {
            "swc_info": ["C", "D", "E", "H", "I", "K"],
            "ib_data": ["C"],
            "ports": ["C", "E", "F", "G"],
            "adt_primitive": ["B", "D", "G", "I"],
            "adt_composite": ["B", "C", "D"],  # Includes Column B for ARRAY check
            "idt": ["B", "C", "D"]  # Ensures "idt" is checked properly
        }
        
        # Ensure "idt" sheet is detected
        if "idt" not in wb.sheetnames:
            print("⚠️ Warning: 'idt' sheet not found in workbook!")
        else:
            print("✅ 'idt' sheet detected. Processing...")
        
        # ✅ Iterate through sheets and check naming rules
        for sheet_name, columns in naming_sheets.items():
            if sheet_name not in wb.sheetnames:
                continue  # Skip if the sheet does not exist
        
            sheet = wb[sheet_name]
        
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                data = {col: row[ord(col.upper()) - 65] if row else "" for col in columns}
            
                # Handle both adt_composite and idt together
                if sheet_name in ["adt_composite", "idt"]:
                    column_b_value = data.get("B", "")
                    column_d_value = data.get("D", "")
                    if column_b_value in ["ARRAY", "ARRAY_FIXED", "ARRAY_VARIABLE"]:
                        if not str(column_d_value).isdigit():
                            errors["Critical"].append(
                                f"[{sheet_name}] Column D must be numeric when Column B is '{column_b_value}' at D{row_idx}: {column_d_value}"
                            )
                    elif column_b_value == "RECORD":
                        if str(column_d_value).isdigit():
                            errors["Critical"].append(
                                f"[{sheet_name}] Column D must NOT be numeric when Column B is 'RECORD' at D{row_idx}: {column_d_value}"
                            )
                elif sheet_name == "ports":
                    column_g_value = data.get("G", "")

                    if column_g_value == "TriggerInterface":
                        if not str(column_g_value).isdigit():
                            errors["Critical"].append(
                                f"[{sheet_name}] Column G must be numeric when Column D is 'TriggerInterface' at G{row_idx}: {column_g_value}"
                            )
                # General Naming Convention Check
                for col in columns:
                    if col == "D" and sheet_name in ["adt_composite", "idt"]:
                        continue 
                    if col == "G" and sheet_name in ["ports"]:
                        continue
            
                    cell_ref = f"{col}{row_idx}"
                    name = data.get(col, "")
            
                    if not re.match(r"^[A-Za-z][A-Za-z0-9_]*$", str(name)):
                        errors["Critical"].append(f"[{sheet_name}] Invalid name format at {cell_ref}: {name}")

        ### 🟡 Duplicate & Definition Consistency Rules ('excel_rule_3') ###
        """"
        excel_rule_3: Duplicate Value Checks
                Rule Definition:
                        This rule mandates the identification and flagging of duplicate values within specified columns of the designated Excel sheets. The rule also incorporates
                        checks for consistency in associated data when duplicates are permitted. Specific Column and Duplicate Validation Requirements:
                1. "swc_info" Sheet:
                        * Columns H, I, and K: Values within these columns must be unique; no duplicates are allowed.
                2. "ib_data" Sheet:
                        * Column C: Values within this column must be unique; no duplicates are allowed.
                3. "ports" Sheet:
                        * Column C: Values within this column must be unique; no duplicates are allowed.
                        * Column E: Duplicate values are permitted. However, when a value in column E is duplicated, the corresponding values in columns D, F, and G (where applicable)
                                must be identical across all instances of the duplicate.
                        * "G" column validation is only applicable where column "D" value is either 'ClientServerInterface', 'TriggerInterface' or 'ModeSwitchInterface'.
                        * Example: If cell E4 is duplicated in cell E11, then cells D4, F4, and G4 (if applicable) must have the same values as cells D11, F11, and G11, respectively.
                        * Column F: Duplicate values are permitted, but only across distinct interfaces (defined by column E). Within a merged range of cells in column E, the
                                corresponding values in column F must be unique.
                        * Example: If cells E6 to E11 are merged, then cells F6 to F11 must contain unique values.
                        * Column G: Duplicate values are permitted, but only across distinct interfaces (defined by column E). Within a merged range of cells in column E, the 
                                corresponding values in column G must be unique if the corresponding column "D" value is either 'ClientServerInterface', 'TriggerInterface' or
                                'ModeSwitchInterface'.
                        * Example: If cells E6 to E11 are merged, then cells G6 to G11 must contain unique values if the corresponding column "D" value is either
                                'ClientServerInterface', 'TriggerInterface' or 'ModeSwitchInterface'.
                4. "adt_primitive" Sheet:
                        * Columns B and G: Values within these columns must be unique; no duplicates are allowed.
                        * Column D: Duplicate values are permitted. However, when a value in column D is duplicated, the corresponding values in columns E, F, G, and H must be 
                                identical across all instances of the duplicate.
                        * Example: If cell D4 is duplicated in cell D11, then cells E4, F4, G4, and H4 must have the same values as cells E11, F11, G11, and H11, respectively.
                        * Column I: Duplicate values are permitted. However, when a value in column I is duplicated, the corresponding values in columns J, K, and L must be 
                                identical across all instances of the duplicate.
                        * Example: If cell I4 is duplicated in cell I11, then cells J4, K4, and L4 must have the same values as cells J11, K11, and L11, respectively.
                5. "adt_composite" Sheet:
                        * Columns C and D: Values within these columns must be unique; no duplicates are allowed, with the exception of numerical values in column D.
                6. "idt" Sheet:
                        * Columns C and D: Values within these columns must be unique; no duplicates are allowed, with the exception of numerical values in column D.
        """
        duplicate_sheets = {
            "swc_info": ["H", "I", "K"],
            "ib_data": ["C"],
            "ports": ["C"],
            "adt_primitive": ["B", "G"],
            "adt_composite": ["C", "D"],  # D has special handling
            "idt": ["C", "D"]  # D has special handling
        }
        for sheet_name, columns in duplicate_sheets.items():
            sheet = wb[sheet_name]
            merged_ranges = get_merged_cells(sheet)  # Get merged cell mappings
            for col in columns:
                seen = set()
                for row_idx, row in enumerate(sheet.iter_rows(
                        min_row=2, min_col=ord(col.upper()) - 64, max_col=ord(col.upper()) - 64, values_only=True), start=2):
                    value = row[0]
                    cell_ref = f"{col}{row_idx}"
                    # Skip empty (None) values entirely
                    if value in [None, ""]:
                        continue  # Do not check empty values for duplication
                    # Handle merged cells (if part of a merged range, log as info and skip)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[{sheet_name}] Merged cell {cell_ref} is expected to have the same value")
                        continue  # Skip checking duplicates for merged empty cells
                    # Special Handling for Column D in `adt_composite` & `idt`
                    if col == "D" and sheet_name in ["adt_composite", "idt"]:
                        if isinstance(value, (int, float)):  # If numerical, duplication is OK (log as Info)
                            if value in seen:
                                errors["Info"].append(f"[{sheet_name}] Duplicate numerical value at {cell_ref}: {value}")
                        else:  # If alphanumeric, duplication is NOT OK (log as Critical Error)
                            if value in seen:
                                errors["Critical"].append(f"[{sheet_name}] Duplicate non-numeric value at {cell_ref}: {value}")
                    else:
                        # General duplicate check for all other columns (log as Critical Error)
                        if value in seen:
                            errors["Critical"].append(f"[{sheet_name}] Duplicate value at {cell_ref}: {value}")
                    seen.add(value)  # Add value to seen set



        ### 🔵 Reference Validation ('excel_rule_4')###

        """"

        excel_rule_4: Referential Integrity
        Rule Definition:
                This rule establishes referential integrity between specified columns across different Excel sheets, ensuring that values in certain columns correspond to valid
                entries in other designated columns. Specific Column and Referential Integrity Requirements:
                        1. "ib_data" Sheet:
                                * Column F: Values in this column must exist within column H of the "swc_info" sheet.
                                * Column D: Values in this column must exist within column B of the "adt_primitive" sheet.
                        2. "ports" Sheet:
                                * Column I: Values in this column must exist within column H of the "swc_info" sheet.
                                * Column H: Values in this column must exist within column B of the "adt_primitive" sheet.
        """
        references = {
            ("ib_data", "F"): ("swc_info", "H"),
            ("ib_data", "D"): ("adt_primitive", "B"),
            ("ports", "I"): ("swc_info", "H"),
            ("ports", "H"): ("adt_primitive", "B")
        }
        for (sheet_name, col), (ref_sheet, ref_col) in references.items():
            if sheet_name not in wb.sheetnames or ref_sheet not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            ref_values = {row[0] for row in wb[ref_sheet].iter_rows(min_row=2, min_col=ord(ref_col)-64, max_col=ord(ref_col)-64, values_only=True)}
            for row_idx, row in enumerate(sheet.iter_rows(
                    min_row=2, min_col=ord(col)-64, max_col=ord(col)-64, values_only=True), start=2):
                value = row[0]
                cell_ref = f"{col}{row_idx}"
                if value not in ref_values:
                    errors["Critical"].append(f"[{sheet_name}] Invalid reference at {cell_ref}: {value} (not in {ref_sheet}.{ref_col})")

        # Rule - 5
        column_enum_mapping = {
            "swc_info": {
                "B": ["ApplicationSwComponentType",
                    "SensorActuatorSwComponentType",
                    "ComplexDeviceDriverSwComponentType",
                    "ParameterSwComponentType",
                    "NvBlockSwComponentType",
                    "ServiceSwComponentType",
                    "EcuAbstractionSwComponentType",
                    "ServiceProxySwComponentType",
                    "CompositionSwComponentType"],
                "F": ["canBeTerminated", "canBeTerminatedAndRestarted", "NO-SUPPORT"],
                "G": ["true", "false"],
                "J": ["false", "true"],
                "L": ["AsynchronousServerCallReturnsEvent",
                    "BackgroundEvent",
                    "DataReceivedEvent",
                    "DataReceiveErrorEvent",
                    "DataSendCompletedEvent",
                    "DataWriteCompletedEvent",
                    "ExternalTriggerOccurredEvent",
                    "InitEvent",
                    "InternalTriggerOccurredEvent",
                    "ModeSwitchedAckEvent",
                    "OperationInvokedEvent",
                    "OsTaskExecutionEvent",
                    "SwcModeManagerErrorEvent",
                    "SwcModeSwitchEvent",
                    "TimingEvent",
                    "TransformerHardErrorEvent"]
            },
            "ib_data": {
                "B": ["SharedParameter", "PerInstanceParameter", "ImplicitInterRunnableVariables",
                    "ExplicitInterRunnableVariable", "PerInstanceMemory", "ArTypedPerInstanceMemory",
                    "StaticMemory", "ConstantMemory"],
                "D": ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
                    "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
                    "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"],
                "G": ["NOT-ACCESSIBLE", "READ", "READ-WRITE"],
                "H": ["CONST", "FIXED", "MEASUREMENT-POINT", "QUEUED", "STANDARD"],
                "I": ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
                    "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
                    "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"]
            },
            "ports": {
                "B": ["ReceiverPort", "SenderPort"],
                "D": ["SenderReceiverInterface", "ParameterInterface", "ClientServerInterface",
                    "NvDataInterface", "ModeSwitchInterface", "TriggerInterface"]
            },
            "adt_primitive": {
                "C": ["BOOLEAN", "Value"],
                "E": ["IDENTICAL", "TEXTTABLE", "LINEAR", "SCALE_LINEAR", "SCALE_LINEAR_AND_TEXTTABLE",
                    "RAT_FUNC", "SCALE_RAT_FUNC", "SCALE_RATIONAL_AND_TEXTTABLE", "TAB_NOINTP",
                    "BITFIELD_TEXTTABLE"],
                "H": ["Ampr", "AmprPerSec", "AmprSec", "Bar", "BarPerSec", "Bel", "Bit", "BitPerSec",
                    "Byte", "BytPerSec", "Cd", "CentiMtr", "CentiMtrSqd", "Coulomb", "Day", "DeciBel",
                    "Deg", "DegCgrd", "DegPerSec", "Frd", "Gr", "GrPerLtr", "GrPerMol", "GrPerSec",
                    "HectoPa", "HectoPaPerSec", "HectoPaPerVolt", "HectoPaSecPerMtrCubd", "Henry",
                    "HenryPerMtr", "Hr", "Hz", "Jou", "JouPerKelvin", "JouPerKiloGr", "JouPerKiloGrPerKelvin",
                    "JouPerMol", "JouPerMolPerKelvin", "Kat", "KelvinAbslt", "KelvinPerSec", "KelvinRel",
                    "KiloBitPerSec", "KiloGr", "KiloGrPerHr", "KiloGrPerMtrCubd", "KiloGrPerSec", "KiloGrSqrMtr",
                    "KiloHz", "KiloJou", "KiloMtr", "KiloMtrPerHr", "KiloMtrPerHrPerSec", "KiloNwt",
                    "KiloNwtMtrPerSec", "KiloOhm", "KiloVolt", "KiloWatt", "KiloWattHr", "KiloWattHrPer100KiloMtr",
                    "Ltr", "LtrPer100KiloMtr", "LtrPerHr", "LtrPerKiloMtr", "MegaBitPerSec", "MegaHz", "MegaJou",
                    "MegaOhm", "MegaPa", "MegaWatt", "MicroAmpr", "MicroFrd", "MicroGr", "MicroJou", "MicroLtr",
                    "MicroLtrPerSec", "MicroMtr", "MicroSec", "MicroTesla", "MilliAmpr", "MilliAmprPerSec",
                    "MilliBar", "MilliFrd", "MilliGr", "MilliGrPerSec", "MilliJou", "MilliLtr", "MilliMtr",
                    "MilliMtrCubd", "MilliMtrCubdPerSec", "MilliOhm", "MilliSec", "MilliTesla", "MilliVolt",
                    "MilliVoltPerSec", "MilliWatt", "Mins", "Ml", "MlPerHr", "Mol", "MolPerLtr", "MolPerLtrPerSec",
                    "MolPerMtrCubd", "MolPerSec", "Mtr", "MtrCubd", "MtrCubdPerHr", "MtrCubdPerHrPerSec", "MtrCubdPerKiloGr",
                    "MtrPerSec", "MtrPerSecCubd", "MtrPerSecSqd", "MtrSqd", "MtrSqdPerSec", "NanoFrd", "NanoSec", "NoUnit",
                    "Nwt", "NwtMtr", "NwtMtrPerRpm", "NwtMtrPerRpmPerSec", "NwtMtrPerSec", "NwtMtrSec", "NwtMtrSqrSec",
                    "NwtPerMtr", "NwtSecPerMtr", "Ohm", "Pa", "PaPerMtrCubdPerSec", "PaPerSec", "PaSec", "Perc",
                    "PercPerMilliSec", "PercPerSec", "PerMille", "PerMin", "PerMtr", "PerSec", "PerSecSqd", "PicoFrd",
                    "PicoSec", "Ppm", "Rad", "RadPerSec", "RadPerSecSqd", "Rpm", "RpmPerSec", "S", "Sec", "SPerMtr",
                    "SPerMtrPerSec", "Tesla", "Tonne", "Volt", "VoltPerMilliSec", "VoltPerMtr", "VoltPerSec", "Watt",
                    "WattPerKelvinPerMtrSqd", "WattPerMtrPerKelvin", "WattPerSec", "WattSec", "Wb", "Yr"],
                "J" : ["physConstrs", "internalConstrs"],
                "M" : ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
                    "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
                    "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"],
            },
            "adt_composite" : {
                "B" : ["RECORD", "ARRAY"],
                "E" : ["APDT", "ARDT", "AADT", "IDT", "FIXED", "VARIABLE"],
                "F" : ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
                "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
                "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"],
            },
            "idt" : {
                "B" : ["ARRAY_FIXED", "ARRAY_VARIABLE", "PRIMITIVE", "RECORD"],
                "E" : ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
                "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
                "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"],
            }
        }

        for sheet_name in wb.sheetnames:
            if sheet_name in column_enum_mapping:
                sheet = wb[sheet_name]
                print(f"\n🔍 Checking Sheet: {sheet_name}")
                
                # Debug: Read and print the first row (column headers)
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                print(f"📌 Column Headers: {headers}")
    
                for col_letter, allowed_values in column_enum_mapping[sheet_name].items():
                    col_index = ord(col_letter) - ord('A') + 1  # Convert column letter to 1-based index
    
                    print(f"🛠 Checking Column {col_letter} (Index: {col_index})")  # Debug info
    
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, values_only=True), start=2):
                        cell_value = row[0]  # Extracting value from the tuple
                        
                        print(f"Row {row_idx}, Col {col_letter}: {cell_value}")  # Debug print
                        if cell_value is not None and cell_value not in allowed_values:
                            errors["Critical"].append(
                                f"Error in {sheet_name}, Cell {col_letter}{row_idx}: '{cell_value}' is invalid.\n"
                                f"  ➝ Allowed values: {', '.join(allowed_values)}"
                            )

        """
            rule no  6 : Number of values allowed limit is one
    
            for following cell there should be one value only, user can not add any other value to it , if added then its critical error
            
                "swc_info":  B, C, D, E,F,G
        """
        multi_value_pattern = re.compile(r"[,*&|/]| and ")

        # Iterate through all sheets
        for sheet_name in wb.sheetnames:
            sheet = wb["swc_info"]

            #defined alloed columns (Excel Column Letters)
            allowed_columns = ["B", "C", "D", "E", "F", "G"]

            col_indices = [ord(col) - ord('A') + 1 for col in allowed_columns]
            # Iterate through all rows and columns
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                for col_idx in col_indices:
                    cell_value = row[col_idx - 1]
                    if cell_value and isinstance(cell_value, str):
                        # Check if cell contains multiple values
                        if multi_value_pattern.search(cell_value):
                            errors["Critical"].append(
                                f"Error in {sheet_name}, Cell {col_letter}{row_idx}: '{cell_value}' contains multiple values."
                            )
        """
            # RULE - 7 -> Correct [MIN, MAX] and check for valid ranged values  
        
        """   
        data_type_ranges = {
            "boolean": (0, 1),
            "ConstVoidPtr": None,  # No range validation needed
            "VoidPtr": None,  # No range validation needed
            "float32": (-3.4e38, 3.4e38),  # IEEE 754 single-precision float range
            "float64": (-1.8e308, 1.8e308),  # IEEE 754 double-precision float range
            "sint8": (-128, 127),
            "sint8_least": (-128, 127),
            "sint16": (-32768, 32767),
            "sint16_least": (-32768, 32767),
            "sint32": (-2147483648, 2147483647),
            "sint32_least": (-2147483648, 2147483647),
            "sint64": (-9223372036854775808, 9223372036854775807),
            "uint8": (0, 255),
            "uint8_least": (0, 255),
            "uint16": (0, 65535),
            "uint16_least": (0, 65535),
            "uint32": (0, 4294967295),
            "uint32_least": (0, 4294967295),
            "uint64": (0, 18446744073709551615),
        }
                        
        sheet = wb["adt_primitive" ]
    
        # Find column indexes dynamically
        col_cons = ord('I') - ord('A') + 1
        col_min = ord('K') - ord('A') + 1
        col_max = ord('L') - ord('A') + 1
        col_data_type = ord('M') - ord('A') + 1
    
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            const_name, min_value, max_value, data_type = row[col_cons - 1], row[col_min - 1], row[col_max - 1], row[col_data_type - 1]
    
            if data_type in data_type_ranges:
                min_allowed, max_allowed = data_type_ranges[data_type]
    
                # Validate min/max within allowed range*\
                
                if min_value is not None and max_value is not None:
                    if not (min_allowed <= min_value <= max_allowed):
                        errors["Critical"].append(f"Error in 'add_primitive' for Data Constraint: {const_name}, Cell K{row_idx}: Min value {min_value} out of range for {data_type}. Allowed: {min_allowed}-{max_allowed}")
                    if not (min_allowed <= max_value <= max_allowed):
                        errors["Critical"].append(f"Error in 'add_primitive' for Data Constraint: {const_name}, Cell L{row_idx}: Max value {max_value} out of range for {data_type}. Allowed: {min_allowed}-{max_allowed}")
                    if min_value > max_value:
                        errors["Critical"].append(f"Error in 'add_primitive' for Data Constraint: {const_name}, Row {row_idx}: Min value {min_value} should not be greater than Max value {max_value}")

        """
            RULE - 8
                if sheet name = "adt_composite"
                    column B == "ARRAY":
                        1. then in corresponding column E only ["FIXED", "VARIABLE"] allowed
                        2. in column F, value must comes from either of them
                            enum_list : ["boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32", "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32", "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"]
                            or values in column B in sheetname = "adt_primitive"
                            or column C in sheetname="idt" if corresponding column B has value "PRIMITIVE"
                    column B == "Record":
                        1. then in corresponding column E only ["APDT", "ARDT", "AADT", "IDT"] allowed
                        2. in column F, value must comes from either of them
                            APDT ( F column value should be from B column value of adt_primitive sheet)
                            or ARDT ( F column value should be from C column value of adt_composite sheet other than current which has corresponding B column value is Record )
                            or AADT (F column value should be from C column value of adt_composite sheet which has corresponding B column value is Array)
                            or IDT (F column value should be either from actual impl data type which are from enum_list sheet from L3 to L21 or custom idt from idt sheet which must other than current Record name (name should come from C column))
        """
        naming_sheets = {
            "adt_composite": ["B", "C", "D", "E", "F"],  # Ensure required columns are present
            "adt_primitive": ["B"],
            "idt": ["B", "C"]
        }
        
        enum_list = [
            "boolean", "ConstVoidPtr", "float32", "float64", "sint16", "sint16_least", "sint32",
            "sint32_least", "sint64", "sint8", "sint8_least", "uint16", "uint16_least", "uint32",
            "uint32_least", "uint64", "uint8", "uint8_least", "VoidPtr"
        ]
        
        # Extract column B values from adt_primitive
        adt_primitive_b_values = set()
        if "adt_primitive" in wb.sheetnames:
            adt_primitive_sheet = wb["adt_primitive"]
            adt_primitive_b_values = {row[0] for row in adt_primitive_sheet.iter_rows(min_row=2, values_only=True) if row[0]}
        
        # Extract column C values from idt where column B = "PRIMITIVE"
        idt_primitive_c_values = set()
        idt_c_values = set()
        if "idt" in wb.sheetnames:
            idt_sheet = wb["idt"]
            idt_primitive_c_values = {row[1] for row in idt_sheet.iter_rows(min_row=2, values_only=True) if row[0] == "PRIMITIVE"}
            idt_c_values = {row[1] for row in idt_sheet.iter_rows(min_row=2, values_only=True)}
        
        # Extract column C values from adt_composite for "RECORD" and "ARRAY"
        adt_composite_record_c_values = set()
        adt_composite_array_c_values = set()
        if "adt_composite" in wb.sheetnames:
            adt_composite_sheet = wb["adt_composite"]
            for row in adt_composite_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == "RECORD":
                    adt_composite_record_c_values.add(row[1])
                elif row[0] == "ARRAY":
                    adt_composite_array_c_values.add(row[1])
        
        for sheet_name, columns in naming_sheets.items():
            if sheet_name not in wb.sheetnames:
                continue
        
            sheet = wb[sheet_name]
            merged_cells = get_merged_cells(sheet)
        
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                data = {col: row[ord(col.upper()) - 65] if row else "" for col in columns}
        
                # Handle merged cells for column B
                for merged_start, merged_group in merged_cells.items():
                    merged_col, merged_row = re.match(r"([A-Z]+)(\d+)", merged_start).groups()
                    if merged_col == "B" and str(row_idx) in [merged_row] + [cell[1:] for cell in merged_group]:
                        data["B"] = sheet[merged_start].value
        
                # RULE - 8 Implementation
                if data.get("B") == "ARRAY":
                    if data.get("E") not in ["FIXED", "VARIABLE"]:
                        errors["Critical"].append(f"[{sheet_name}] Column E must be 'FIXED' or 'VARIABLE' at E{row_idx}: {data.get('E')}")
                    if data.get("F") not in enum_list and data.get("F") not in adt_primitive_b_values and data.get("F") not in idt_primitive_c_values:
                        errors["Critical"].append(f"[{sheet_name}] Column F must be from enum_list, adt_primitive (B column), or idt (C column when B='PRIMITIVE') at F{row_idx}: {data.get('F')}")
        
                if sheet_name == "adt_composite" and data.get("B") == "RECORD":
                    if data.get("E") not in ["APDT", "ARDT", "AADT", "IDT"]:
                        errors["Critical"].append(f"[{sheet_name}] Column E must be 'APDT', 'ARDT', 'AADT', or 'IDT' at E{row_idx}: {data.get('E')}")
                    
                    if data.get("E") == "APDT" and data.get("F") not in adt_primitive_b_values:
                        errors["Critical"].append(f"[{sheet_name}] Column F must be from adt_primitive (B column) when Column E is APDT at F{row_idx}: {data.get('F')}")
                    
                    if data.get("E") == "ARDT":
                        valid_ardt_values = adt_composite_record_c_values - {data.get("F")}
                        if data.get("F") not in valid_ardt_values:
                            errors["Critical"].append(f"[{sheet_name}] Column F must be from adt_composite (C column where B is RECORD) when Column E is ARDT, excluding itself at F{row_idx}: {data.get('F')}")
                    
                    if data.get("E") == "AADT":
                        valid_aadt_values = adt_composite_array_c_values - {data.get("F")}
                        if data.get("F") not in valid_aadt_values:
                            errors["Critical"].append(f"[{sheet_name}] Column F must be from adt_composite (C column where B is ARRAY) Column E is AADT, excluding itself at F{row_idx}: {data.get('F')}")
                    
                    if data.get("E") == "IDT":
                        valid_idt_values = enum_list + list(idt_c_values - {data.get("F")})
                        if data.get("F") not in valid_idt_values:
                            errors["Critical"].append(f"[{sheet_name}] Column F must be from enum_list or idt (C column), excluding itself at F{row_idx}: {data.get('F')}")
                
        # Print Errors if any
        if errors["Critical"]:
            print("\n".join(errors["Critical"]))
        else:
            print("✅ Rule 8 Validation Passed! No errors found.")

        # ✅ Runnable Access Rule for "ports" sheet (G column validation)
        """
            G column value should be {dra, drpa, drpv) when corresponding D column value either SenderReceiverInterface or NvDataInterface and B column value ReceiverPort
            G column value should be {dsp , dwa) when corresponding D column value either SenderReceiverInterface or NvDataInterface and B column value SenderPort
            G column value must be empty when corresponding D column value ParameterInterface
        """
        if "ports" in wb.sheetnames:
            sheet = wb["ports"]
        
            # Step 1: Store merged cell values properly
            merged_values = {}
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                first_cell_value = sheet.cell(row=min_row, column=min_col).value  # Get the first cell's value
                
                # Assign this value to all merged cells
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_values[(row, col)] = first_cell_value
        
            # Step 2: Iterate over rows and validate rules
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                column_b_value = row[1]  # Column B (Port Type)
                column_d_value = row[3] if row[3] else merged_values.get((row_idx, 4), "")  # Handle merged cells in Column D
                column_g_value = str(row[6]).strip().lower() if row[6] else ""  # Column G (Lowercase and stripped)
        
                # Validation logic
                if column_d_value in ["SenderReceiverInterface", "NvDataInterface"]:
                    if column_b_value == "ReceiverPort" and column_g_value not in {"dra", "drpa", "drpv"}:
                        errors["Critical"].append(
                            f"[ports] Column G must be one of {{dra, drpa, drpv}} when Column D is '{column_d_value}' and Column B is 'ReceiverPort' at G{row_idx}: {column_g_value}"
                        )
                    if column_b_value == "SenderPort" and column_g_value not in {"dsp", "dwa"}:
                        errors["Critical"].append(
                            f"[ports] Column G must be one of {{dsp, dwa}} when Column D is '{column_d_value}' and Column B is 'SenderPort' at G{row_idx}: {column_g_value}"
                        )
        
                if column_d_value == "ParameterInterface" and column_g_value:
                    errors["Critical"].append(
                        f"[ports] Column G must be empty when Column D is 'ParameterInterface' at G{row_idx}: {column_g_value}"
                    )
        
            # Step 3: Print errors
            for error in errors["Critical"]:
                print(error)


        # ✅ Merge Rule Validation
        merge_rules = {
            "swc_info": ["B", "C", "D", "E", "F", "G"],
            "ports": ["B", "C", "D", "E", "F"],
            "adt_primitive": ["B", "C", "D", "E", "H", "I", "J", "K", "L", "M"],
            "adt_composite": ["B", "C"],
            "idt": ["B", "C"]
            }
        
        for sheet_name, columns in merge_rules.items():
            if sheet_name not in wb.sheetnames:
                continue  # Skip if sheet doesn't exist
            sheet = wb[sheet_name]
        
            for merged_cell in sheet.merged_cells.ranges:
                start_row, end_row = merged_cell.min_row, merged_cell.max_row
                start_col, end_col = merged_cell.min_col, merged_cell.max_col
        
                # Convert column indices to column letters (to compare against merge_rules)
                start_col_letter = chr(64 + start_col)
                end_col_letter = chr(64 + end_col)
        
                # Check if any merged columns are outside the allowed ones
                if not any(col in columns for col in [start_col_letter, end_col_letter]):
                    errors["Warning"].append(
                        f"[{sheet_name}] Merge found in {start_col_letter}{start_row}:{end_col_letter}{end_row}, which is not allowed"
                    )
                else:
                    # Additional check based on special conditions (e.g., for "adt_primitive" and "identical" values)
                    if sheet_name == "adt_primitive" and sheet[f"E{start_row}"].value == "identical":
                        continue  # Skip if E column value is 'identical'
        
                    if sheet_name in ["adt_composite", "idt"] and sheet[f"C{start_row}"].value != "Record":
                        continue  # Skip if C column value is NOT "Record"
                        
                    # Append warning if merge is found in allowed columns but needs additional checks
                    errors["Warning"].append(
                        f"[{sheet_name}] Merge found in {start_col_letter}{start_row}:{end_col_letter}{end_row}, which may not be allowed"
                    )

        # ✅ Data Type Reference Rule for "ib_data" (I column) & "ports" (H column)
        
        # Collect reference values from adt_primitive, adt_composite, idt, and enum_list
        valid_values = set()
        
        # Extract B column values from adt_primitive
        if "adt_primitive" in wb.sheetnames:
            sheet = wb["adt_primitive"]
            valid_values.update(cell.value for cell in sheet["B"][1:] if cell.value)  # Skip header
        
        # Extract C column values from adt_composite
        if "adt_composite" in wb.sheetnames:
            sheet = wb["adt_composite"]
            valid_values.update(cell.value for cell in sheet["C"][1:] if cell.value)
        
        # Extract C column values from idt
        if "idt" in wb.sheetnames:
            sheet = wb["idt"]
            valid_values.update(cell.value for cell in sheet["C"][1:] if cell.value)
        
        # Extract L3 to L21 values from enum_list
        if "enum_list" in wb.sheetnames:
            sheet = wb["enum_list"]
            for row in sheet.iter_rows(min_row=3, max_row=21, min_col=12, max_col=12, values_only=True):
                if row[0]:
                    valid_values.add(row[0])
        
        # Validate "I" column in "ib_data"
        if "ib_data" in wb.sheetnames:
            sheet = wb["ib_data"]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                column_i_value = row[8]  # I column (Index 8)
                if column_i_value and column_i_value not in valid_values:
                    errors["Critical"].append(
                        f"[ib_data] Invalid reference in Column I at I{row_idx}: {column_i_value}"
                    )
        
        # Validate "H" column in "ports"
        if "ports" in wb.sheetnames:
            sheet = wb["ports"]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                column_h_value = row[7]  # H column (Index 7)
                if column_h_value and column_h_value not in valid_values:
                    errors["Critical"].append(
                        f"[ports] Invalid reference in Column H at H{row_idx}: {column_h_value}"
                    )

    except Exception as e:
        errors["Critical"].append(f"Error reading Excel file: {str(e)}")
    return errors


def print_colored_errors(errors):

   """ Prints errors in color-coded format. """

   for severity, msgs in errors.items():

       if msgs:

           color = RED if severity == "Critical" else (YELLOW if severity == "Warning" else BLUE)

           for msg in msgs:

               print(f"{color}{msg}{RESET}")  # Apply color and reset after message

def log_errors(errors, attempt_number):
   """ Logs validation errors with severity levels and timestamps. """
   timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   with open("validation_log.txt", "a", encoding="utf-8") as log_file:
       log_file.write(f"\n=== Validation Attempt {attempt_number} at {timestamp} ===\n")
       for severity, msgs in errors.items():
           if msgs:
               log_file.write(f"\n[{severity} ERRORS]\n")
               for msg in msgs:
                   log_file.write(f"{msg}\n")
   print(f"{YELLOW}Errors logged in 'validation_log.txt'. Please fix them before retrying.{RESET}")

def generate_summary(initial_errors, final_errors, attempts):

   """ Generates a validation summary report. """

   fixed_errors = len(initial_errors["Critical"]) - len(final_errors["Critical"])

   summary_report = f"""

   ============================

   Validation Summary Report

   ============================

   📅 Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

   🔄 Validation Attempts: {attempts}

   ❌ Initial Critical Errors: {len(initial_errors["Critical"])}

   ⚠️ Initial Warnings: {len(initial_errors["Warning"])}

   ✅ Critical Errors Fixed: {fixed_errors}

   """

   print(summary_report)

   with open("validation_summary.txt", "w", encoding="utf-8") as summary_file:

       summary_file.write(summary_report)

   print(f"{GREEN}✔ Summary saved in 'validation_summary.txt'.{RESET}") 

def generate_html_report(errors, attempts):
    """
    Generates an HTML report for validation errors.
    """
    report_filename = f"validation_report_attempt_{attempts}.html"
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Validation Report - Attempt {attempts}</title>
    <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1, h2 {{ color: #333; }}
            .critical {{ color: red; font-weight: bold; }}
            .warning {{ color: orange; font-weight: bold; }}
            .info {{ color: blue; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
    </style>
    </head>
    <body>
    <h1>Excel Validation Report - Attempt {attempts}</h1>
    """
    # Add Critical Errors Section
    if errors["Critical"]:
        html_content += "<h2 class='critical'>❌ Critical Errors</h2><ul>"
        for error in errors["Critical"]:
            html_content += f"<li class='critical'>{error}</li>"
        html_content += "</ul>"
    # Add Warnings Section
    if errors["Warning"]:
        html_content += "<h2 class='warning'>⚠️ Warnings</h2><ul>"
        for error in errors["Warning"]:
            html_content += f"<li class='warning'>{error}</li>"
        html_content += "</ul>"
    # Add Info Section
    if errors["Info"]:
        html_content += "<h2 class='info'>ℹ️ Info Messages</h2><ul>"
        for error in errors["Info"]:
            html_content += f"<li class='info'>{error}</li>"
        html_content += "</ul>"
    # If no errors exist
    if not any(errors.values()):
        html_content += "<h2>✅ No validation errors found.</h2>"
    html_content += "</body></html>"
    # Write the HTML content to a file
    with open(report_filename, "w", encoding="utf-8") as file:
        file.write(html_content)
    print(f"\n📄 HTML Report Generated: {report_filename}")   
